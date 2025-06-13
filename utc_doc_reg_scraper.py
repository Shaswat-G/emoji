# -----------------------------------------------------------------------------
# Script: utc_doc_reg_scraper.py
# Summary: Scrapes and consolidates UTC document register metadata from
#          Unicode.org for multiple years into structured CSV and Excel files.
# Inputs:  None (fetches HTML from Unicode.org for 2011–2021)
# Outputs: utc_register_{year}.csv (per year), utc_register_all.xlsx (combined)
# Context: Part of a research pipeline analyzing UTC's emoji proposal and
#          decision-making processes using public data.
# -----------------------------------------------------------------------------

import requests
from bs4 import BeautifulSoup
import csv
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

def extract_document_info(soup):
    tables = soup.find_all('table', class_='subtle')
    if not tables:
        return []

    main_table = None
    for table in tables:
        headers = [th.get_text().strip() for th in table.find_all('th')]
        if (('Doc №' in headers) or ('Doc Number' in headers)) and 'Subject' in headers and 'Source' in headers and 'Date' in headers:
            main_table = table
            break
    
    if not main_table:
        return []
    
    documents = []
    for row in main_table.find_all('tr')[1:]:  # Skip header row
        cells = row.find_all('td')
        if len(cells) >= 4:
            # Extract document number and link
            doc_cell = cells[0]
            doc_link = doc_cell.find('a')
            doc_num = doc_cell.get_text().strip()
            doc_url = doc_link['href'] if doc_link else ""
            subject = cells[1].get_text().strip()
            source = cells[2].get_text().strip()
            date = cells[3].get_text().strip()
            
            documents.append({
                'doc_num': doc_num,
                'doc_url': doc_url,
                'subject': subject,
                'source': source,
                'date': date
            })
    
    return documents

def scrape_utc_register(year, save_to_csv=True):
    url = f"https://www.unicode.org/L2/L{year}/Register-{year}.html"
    print(f"Scraping {url}...")
    
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        
        documents = extract_document_info(soup)
        
        if save_to_csv and documents:
            output_file = f"utc_register_{year}.csv"
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['doc_num', 'doc_url', 'subject', 'source', 'date']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                for doc in documents:
                    writer.writerow(doc)
            print(f"Saved {len(documents)} documents to {output_file}")
        
        return documents
    
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        return []

def main():
    years = range(2011, 2022)
    
    all_documents = []
    for year in years:
        documents = scrape_utc_register(year, save_to_csv=False)
        if documents:
            all_documents.extend(documents)
            print(f"Found {len(documents)} documents for {year}")
    
    print(f"Total documents scraped: {len(all_documents)}")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "UTC_Register_Docs"

    headers = ['doc_num', 'doc_url', 'subject', 'source', 'date']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Write data
    for row_num, doc in enumerate(all_documents, 2):
        sheet[f"A{row_num}"] = doc['doc_num']
        sheet[f"B{row_num}"] = doc['doc_url']
        sheet[f"C{row_num}"] = doc['subject']
        sheet[f"D{row_num}"] = doc['source']
        sheet[f"E{row_num}"] = doc['date']

    workbook.save("utc_register_all.xlsx")
    print(f"Saved all {len(all_documents)} documents to utc_register_all.xlsx")

if __name__ == "__main__":
    main()
